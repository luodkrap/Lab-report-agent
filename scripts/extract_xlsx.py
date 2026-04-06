#!/usr/bin/env python3
"""xlsx 파일에서 텍스트를 추출하여 tmp/extracted_data.txt에 저장하는 스크립트."""

import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다. 설치 중...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


def extract_xlsx(file_path: str, output_path: str) -> None:
    """xlsx 파일의 모든 시트를 탭 구분 텍스트로 변환하여 저장한다."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"[시트명: {sheet_name}]")

        for row in ws.iter_rows():
            cell_values = []
            for cell in row:
                if cell.value is None:
                    cell_values.append("")
                else:
                    cell_values.append(str(cell.value))
            lines.append("\t".join(cell_values))

        lines.append("")  # 시트 간 빈 줄

    wb.close()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"추출 완료: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python scripts/extract_xlsx.py <xlsx 파일 경로>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"파일을 찾을 수 없습니다: {xlsx_path}")
        sys.exit(1)

    project_root = Path(__file__).resolve().parent.parent
    output_file = str(project_root / "tmp" / "extracted_data.txt")

    extract_xlsx(xlsx_path, output_file)
