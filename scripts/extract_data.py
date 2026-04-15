#!/usr/bin/env python3
"""xlsx와 csv 파일을 통합 추출하여 tmp/extracted_data.txt에 저장하는 스크립트.

사용법:
    python3 scripts/extract_data.py <파일 또는 디렉터리>

디렉터리를 지정하면 하위의 모든 .xlsx, .csv 파일을 수집해
tmp/extracted_data.txt에 단일 텍스트로 직렬화한다.
"""

from __future__ import annotations

import csv
import os
import sys
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다. 설치 중...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


SUPPORTED_EXTS = {".xlsx", ".csv"}


def collect_files(target: Path) -> list[Path]:
    if target.is_file():
        if target.suffix.lower() not in SUPPORTED_EXTS:
            raise ValueError(f"지원하지 않는 파일 형식: {target.suffix}")
        return [target]
    if target.is_dir():
        files: list[Path] = []
        for ext in SUPPORTED_EXTS:
            files.extend(sorted(target.rglob(f"*{ext}")))
        return files
    raise FileNotFoundError(f"경로를 찾을 수 없습니다: {target}")


def serialize_xlsx(file_path: Path) -> list[str]:
    lines: list[str] = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"[시트명: {sheet_name}]")
        for row in ws.iter_rows():
            cell_values = ["" if cell.value is None else str(cell.value) for cell in row]
            lines.append("\t".join(cell_values))
        lines.append("")
    wb.close()
    return lines


def serialize_csv(file_path: Path) -> list[str]:
    encodings = ("utf-8-sig", "utf-8", "cp949")
    last_error: Exception | None = None
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
            lines: list[str] = []
            for row in rows:
                lines.append("\t".join(row))
            lines.append("")
            return lines
        except UnicodeDecodeError as e:
            last_error = e
            continue
    raise RuntimeError(f"CSV 디코딩 실패 ({file_path}): {last_error}")


def serialize_file(file_path: Path) -> list[str]:
    header = f"=== 파일: {file_path.name} ==="
    if file_path.suffix.lower() == ".xlsx":
        body = serialize_xlsx(file_path)
    elif file_path.suffix.lower() == ".csv":
        body = [f"[CSV: {file_path.name}]", *serialize_csv(file_path)]
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {file_path.suffix}")
    return [header, *body, ""]


def write_output(blocks: Iterable[list[str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        for block in blocks:
            f.write("\n".join(block))
            f.write("\n")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("사용법: python3 scripts/extract_data.py <파일 또는 디렉터리>")
        return 1

    target = Path(argv[1]).resolve()
    project_root = Path(__file__).resolve().parent.parent
    output_path = project_root / "tmp" / "extracted_data.txt"

    try:
        files = collect_files(target)
    except (FileNotFoundError, ValueError) as e:
        print(f"오류: {e}")
        return 1

    if not files:
        print(f"처리할 .xlsx 또는 .csv 파일이 없습니다: {target}")
        write_output([], output_path)
        return 0

    blocks: list[list[str]] = []
    for file_path in files:
        try:
            blocks.append(serialize_file(file_path))
            print(f"처리: {file_path}")
        except Exception as e:
            print(f"건너뜀 ({file_path}): {e}")

    write_output(blocks, output_path)
    print(f"추출 완료: {output_path} (파일 {len(blocks)}개)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
